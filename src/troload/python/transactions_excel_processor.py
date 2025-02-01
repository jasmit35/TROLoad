"""
transactions_excel_processor.py

This module provides the TransactionsExcelProcessor class for processing
transactions from an Excel spreadsheet.

The TransactionsExcelProcessor class is used to process transactions from an
Excel spreadsheet. It reads the transactions from the spreadsheet and inserts
them into the database.
"""

import datetime

import openpyxl
from accounts import AccountsTable
from categories_table import CategoriesTable
from category_data import CategoryData
from std_logging import function_logger, getLogger
from transactions import InvalidTransactionException, Transaction, TransactionsTable

trans_date_col = 1
account_col = 2
number_col = 3
description_col = 4
memo_col = 5
category_col = 6
tag_col = 7
tax_col = 8
cleared_col = 9
amount_col = 10


# ======================================================================
class TransactionsExcelProcessor:
    def __init__(self, db_conn, file_path, report):
        self._logger = getLogger()
        self._logger.info(f"Begin 'TransactionsExcelProcessor.__init__({file_path=})")

        self._db_conn = db_conn
        self._file_path = file_path
        self._report = report

        self._workbook = openpyxl.load_workbook(filename=file_path)
        self._accounts_table = AccountsTable(db_conn)
        self._categories_table = CategoriesTable(db_conn)
        self._end_of_transactions_label = None

    # ---------------------------------------------------------------------------------------------------------------------
    def __str__(self):
        return f"TransactionsExcelProcessor({self._file_path=})"

    __repr__ = __str__

    # ---------------------------------------------------------------------------------------------------------------------
    def report(self, msg):
        self._report.report(msg)

    # ----------------------------------------------------------------------
    @function_logger
    def process_excel_file(self):
        """
        Process the Excel file by loading new accounts and categories and
        then loading the transactions.
        """
        start_date, end_date = self.extract_date_range()
        self.load_any_new_accounts()
        self.load_any_new_categories()
        self.load_transactions_from_workbook()
        return 0

    # ----------------------------------------------------------------------
    @function_logger
    def extract_date_range(self):
        sheet = self._workbook.active
        for transaction in sheet.iter_rows(
            min_row=3,
            max_row=3,
            min_col=1,
            max_col=1,
            values_only=True,
        ):
            date_label = next(iter(transaction)).split()
            start_date = date_label[3]
            end_date = date_label[5]
            self._end_of_transactions_label = f"{start_date} - {end_date}"
        return start_date, end_date

    #  ----------------------------------------------------------------------
    @function_logger
    def validate_transaction(self, trans):
        """
        Checks a transaction record to make sure certain fields are valid,
        otherwise raise an exception.

        Args:
            trans (_type_): _description_

        Raises:
            InvalidTransactionException:
        """

        date_str = str(trans[trans_date_col])

        if "BALANCE" in date_str:
            raise InvalidTransactionException(trans, f"This transaction does not have a valid date - {date_str}")
        if "Date" in date_str:
            raise InvalidTransactionException(trans, f"This transaction does not have a valid date - {date_str}")

        amount = trans[amount_col]
        if amount is None:
            raise InvalidTransactionException(trans, 'This transaction has an invalid amount - "None"')

    # ----------------------------------------------------------------------
    @function_logger
    def load_any_new_accounts(self):
        accounts_table = AccountsTable(self._db_conn)
        sheet = self._workbook.active

        self.report("\n\n    The following new accounts have been added:\n")

        for transaction in sheet.iter_rows(
            min_row=7,
            max_row=99999,
            min_col=3,
            max_col=3,
            values_only=True,
        ):
            account_name = transaction[0]

            if account_name:
                account_id = accounts_table.get_id(account_name, False)
                if account_id is None:
                    account_id = accounts_table.insert_name(account_name)
                    accounts_table.existing_accounts[account_name] = account_id
                    self.report(f"      {account_id}  {account_name}\n")

    #  ----------------------------------------------------------------------
    @function_logger
    def load_any_new_categories(self):
        sheet = self._workbook.active

        self.report("\n\n    The following new categories have been added:\n")

        for transaction in sheet.iter_rows(
            min_row=6,
            max_row=99999,
            min_col=7,
            max_col=7,
            values_only=True,
        ):
            if len(transaction) < 6:
                continue

            category_name = self.massage_category_name(transaction)

            if category_name:
                cat_id = self._categories_table.select_id_using_name(category_name)
                if cat_id is None:
                    self.report(f"      {category_name}\n")
                    new_category = CategoryData(category_name)
                    self._categories_table.insert(new_category)

    #  ----------------------------------------------------------------------
    @function_logger
    def massage_category_name(self, transaction):
        category_name = transaction[category_col]
        if category_name is None:
            if transaction[number_col] == "Added":
                category_name = "Added"
            if transaction[number_col] == "Bought":
                category_name = "Bought"
            if transaction[number_col] == "Removed":
                category_name = "Removed"
            if transaction[number_col] == "StkSplit":
                category_name = "Stock Split"
            if transaction[number_col] == "Sold":
                category_name = "Sold"
        return category_name

    #  ----------------------------------------------------------------------
    @function_logger
    def load_transactions_from_workbook(self):
        """
        Read each row of the spreadsheet and turn it into a transaction
        record.

        Returns:
            _type_: _description_
        """
        accounts_table = AccountsTable(self._db_conn)
        trans_tab = TransactionsTable(self._db_conn)
        sheet = self._workbook.active

        #  The spreadsheet we are loading from does not repeat all transactions for split transactions.
        #  Theirfore it is necessary to retain the previous transactions.
        previous_transaction_date_string = "1960-01-12 00:00:00"
        previous_account_id = 0
        previous_account_name = ""
        previous_description = ""

        self.report("\n\n    The following transactions have been added:\n")

        rc = 0

        for transaction in sheet.iter_rows(min_row=1, max_row=99999, min_col=1, max_col=11, values_only=True):
            if transaction[1] == self._end_of_transactions_label:
                break

            try:
                self.validate_transaction(transaction)
                account_name = transaction[account_col]
                if account_name:
                    account_id = accounts_table.get_id(account_name)
                    previous_account_id = account_id
                    previous_account_name = account_name
                else:
                    account_id = previous_account_id
                    account_name = previous_account_name

                transaction_date_string = str(transaction[trans_date_col])
                if transaction_date_string == "None":
                    transaction_date_string = previous_transaction_date_string
                else:
                    previous_transaction_date_string = transaction_date_string
                iso_date_string = str(transaction_date_string.split()[0])
                transaction_date = datetime.date.fromisoformat(iso_date_string)

                # category_name = transaction[category_col]
                category_id = self.resolve_category_id(transaction)

                amount = transaction[amount_col]

                this_trans = Transaction(account_id, transaction_date, category_id, amount)

                this_trans.cleared = transaction[cleared_col]
                this_trans.number = transaction[number_col]
                this_trans.tag = transaction[tag_col]

                this_trans.description = transaction[description_col]
                if this_trans.description:
                    previous_description = this_trans.description
                else:
                    this_trans.description = previous_description

                this_trans.memo = transaction[memo_col]
                this_trans.tax_item = transaction[tax_col]

                trans_tab.insert_transaction(this_trans)
                self.report(f"      {account_name}, {transaction_date}, {category_name}, {amount}\n")

            except InvalidTransactionException as e:
                self.report(f"\nError! {e.message}\n")
                self.report(f"{transaction}\n")

        return rc

    #  ----------------------------------------------------------------------
    @function_logger
    def resolve_category_id(self, transaction):
        category_name = self.massage_category_name(transaction[category_col])

        category_id = self._categories_table.select_id_using_name(category_name)

        if category_id is None:
            raise InvalidTransactionException(
                transaction,
                f"This transaction has an invalid category - '{category_name}'",
            )

        return category_id
