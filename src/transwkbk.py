"""
transwkbk.py - Excel workbook of transaction produced by Quicken.
"""
import datetime
import os
import sys

import openpyxl

p = os.path.abspath("../TRO/local/python")
sys.path.insert(1, p)
from accounts import AccountsTable
from categories import CategoriesTable
from transactions import Transaction, TransactionsTable, InvalidTransactionException

trans_date_col = 1
account_col = 2
number_col = 3
description_col = 4
memo_col = 5
category_col = 6
tag_col = 7
tax_col = 8
cleared_col = 9
amount_col = 10


class TransactionWorkbook:
    def __init__(self, this_app, file_name) -> None:
        self.this_app = this_app
        self.workbook = openpyxl.load_workbook(filename=file_name)
        self.accounts_table = AccountsTable(this_app.db_conn)
        self.categories_table = CategoriesTable(this_app.db_conn)
        self.end_of_transactions_label = None

    #  ----------------------------------------------------------------------
    def get_transaction_date_range(self):
        self.this_app.info("begin get_transaction_date_range()")
        sheet = self.workbook.active
        for transaction in sheet.iter_rows(
            min_row=3,
            max_row=3,
            min_col=1,
            max_col=1,
            values_only=True,
        ):
            date_label = list(transaction)[0].split()
            start_date = date_label[3]
            end_date = date_label[5]
            self.end_of_transactions_label = f"{start_date} - {end_date}"
        self.this_app.debug(f"end  get_transaction_date_range - returns {start_date=}, {end_date=}")
        return start_date, end_date

    #  ----------------------------------------------------------------------
    def load_new_accounts_from_workbook(self):

        accounts_table = AccountsTable(self.this_app.db_conn)
        sheet = self.workbook.active

        self.this_app.output("\n\n    The following new accounts have been added:\n")

        for transaction in sheet.iter_rows(
            min_row=7,
            max_row=999,
            min_col=3,
            max_col=3,
            values_only=True,
        ):
            account_name = transaction[0]

            if account_name:
                account_id = accounts_table.get_id(account_name, False)
                if account_id is None:
                    account_id = accounts_table.insert_name(account_name)
                    accounts_table.existing_accounts[account_name] = account_id
                    self.this_app.output(f"      {account_id}  {account_name}\n")

    #  ----------------------------------------------------------------------
    def load_new_categories_from_workbook(self):

        sheet = self.workbook.active

        self.this_app.output("\n\n    The following new categories have been added:\n")

        for transaction in sheet.iter_rows(
            min_row=6,
            max_row=9999,
            min_col=7,
            max_col=7,
            values_only=True,
        ):

            cat_name = transaction[0]

            if cat_name:
                cat_id = self.categories_table.get_id(cat_name, False)
                if cat_id is None:
                    self.this_app.output(f"      {cat_name}\n")
                    cat_id = self.categories_table.insert_category(cat_name)
                    self.categories_table.buffered_categories[cat_name] = cat_id

    #  ----------------------------------------------------------------------
    #  Check that the required fields are present, if not abort
    def validate_transaction(self, trans):

        amount = trans[amount_col]
        if amount is None:
            raise InvalidTransactionException(trans, 'This transaction has an invalid amount - "None"')

    #  ----------------------------------------------------------------------
    def load_transactions_from_workbook(self):
        accounts_table = AccountsTable(self.this_app.db_conn)
        trans_tab = TransactionsTable(self.this_app.db_conn)
        sheet = self.workbook.active

        #  The spreadsheet we are loading from does not repeat all transactions for split transactions.
        #  Theirfore it is necessary to retain the previous transactions.
        previous_transaction_date_string = "1960-01-12 00:00:00"
        previous_account_id = 0
        previous_account_name = ""
        previous_description = ""

        self.this_app.output("\n\n    The following transactions have been added:\n")

        rc = 0

        for transaction in sheet.iter_rows(min_row=9, max_row=9999, min_col=1, max_col=11, values_only=True):
            if transaction[1] == self.end_of_transactions_label:
                break

            try:
                self.validate_transaction(transaction)
                account_name = transaction[account_col]
                if account_name:
                    account_id = accounts_table.get_id(account_name)
                    previous_account_id = account_id
                    previous_account_name = account_name
                else:
                    account_id = previous_account_id
                    account_name = previous_account_name

                transaction_date_string = str(transaction[trans_date_col])
                if transaction_date_string == "None":
                    transaction_date_string = previous_transaction_date_string
                else:
                    previous_transaction_date_string = transaction_date_string
                iso_date_string = str(transaction_date_string.split()[0])
                transaction_date = datetime.date.fromisoformat(iso_date_string)

                category_name = transaction[category_col]
                category_id = self.get_category_id(transaction)

                amount = transaction[amount_col]

                this_trans = Transaction(account_id, transaction_date, category_id, amount)

                this_trans.cleared = transaction[cleared_col]
                this_trans.number = transaction[number_col]
                this_trans.tag = transaction[tag_col]

                this_trans.description = transaction[description_col]
                if this_trans.description:
                    previous_description = this_trans.description
                else:
                    this_trans.description = previous_description

                this_trans.memo = transaction[memo_col]
                this_trans.tax_item = transaction[tax_col]

                trans_tab.insert_transaction(this_trans)
                self.this_app.output(f"      {account_name}, {transaction_date}, {category_name}, {amount}\n")

            except InvalidTransactionException as e:
                self.this_app.error("An unhandable exception has occured. Please review the log file.")
                self.this_app.output(f"\nError! {e.message}\n")
                self.this_app.output(f"{transaction}\n")
                self.this_app.output("\nProcessing is terminating on this file.")
                rc = 16
                break

        return rc

    #  ----------------------------------------------------------------------
    def get_category_id(self, transaction):
        category_name = transaction[category_col]
        if category_name is None:
            if transaction[number_col] == "Added":
                category_name = "Added"
            if transaction[number_col] == "Bought":
                category_name = "Bought"
            if transaction[number_col] == "Removed":
                category_name = "Removed"

        category_id = self.categories_table.get_id(category_name, False)

        if category_id is None:
            raise InvalidTransactionException(
                transaction, f"This transaction has an invalid category - '{category_name}'"
            )

        return category_id
