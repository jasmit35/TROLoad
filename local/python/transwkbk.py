import datetime
import openpyxl

import accounts
import transactions

from categories import CategoriesTable
from transactions import Transaction, TransactionsTable


class TransactionWorkbook():
    def __init__(self, file_name, log, rpt, db_conn) -> None:
        self.workbook = openpyxl.load_workbook(filename=file_name)
        self.log = log
        self.rpt = rpt
        self.db_conn = db_conn

    def get_transaction_date_range(self):
        self.log.debug('begin get_transaction_date_range()')
        sheet = self.workbook.active
        for transaction in sheet.iter_rows(
            min_row=3, max_row=3,
            min_col=1, max_col=1,
            values_only=True,
        ):
            date_label = list(transaction)[0].split()
            start_date = date_label[0]
            end_date = date_label[2]
        return start_date, end_date

    def load_new_accounts_from_workbook(self):
        known_accounts = accounts.select_all_accounts(self.db_conn)

        self.rpt.write("  The following new accounts have been added:\n")

        sheet = self.workbook.active
        for transaction in sheet.iter_rows(
            min_row=7, max_row=999,
            min_col=3, max_col=3,
            values_only=True,
        ):
            account_name = transaction[0]
            if account_name is None:
                continue
            if account_name in known_accounts.values():
                continue
            accounts.add_new_account(self.db_conn, account_name)
            self.rpt.write(f"    {account_name}\n")
        self.log.debug('end   load_new_accounts_from_workbook - returns None')

    def load_new_categories_from_workbook(self):
        categories_dict = CategoriesTable(self.db_conn)

        sheet = self.workbook.active

        self.rpt.write("\n\n  The following new categrories have been added:\n")

        for transaction in sheet.iter_rows(
            min_row=5, max_row=999,
            min_col=7, max_col=7,
            values_only=True,
        ):
            cat_name = transaction[0]

            if cat_name:
                cat_id = categories_dict.get_id(cat_name)
                if cat_id == 0:
                    categories_dict.add_category(cat_name)
                    self.rpt.write(f"    {cat_name}\n")

    def invalid_trans(self, trans):
        category = trans[5]
        if category is None:
            return True

        amount = trans[10]
        if amount is None:
            return True

        return False

    def load_transactions_from_workbook(self):
        categories_dict = CategoriesTable(self.db_conn)
        trans_tab = TransactionsTable(self.db_conn)
        sheet = self.workbook.active

        #  The spreadsheet we are loading from does not repeat all transactions for split transactions.
        #  Theirfore it is necessary to retain the previous transactions.
        previous_transaction_date_string = "1960-01-12 00:00:00"
        previous_account_id = 0
        previous_description = ""

        self.rpt.write("\n\n  The following transactions have been added:\n")

        for transaction in sheet.iter_rows(
            min_row=8, max_row=999,
            min_col=2, max_col=12,
            values_only=True
        ):
            if self.invalid_trans(transaction):
                continue

            account_name = transaction[1]
            if account_name:
                account_id = accounts.get_account_id(account_name)
                previous_account_id = account_id
            else:
                account_id = previous_account_id

            transaction_date_string = str(transaction[0])
            if transaction_date_string == 'None':
                transaction_date_string = previous_transaction_date_string
            else:
                previous_transaction_date_string = transaction_date_string
            iso_date_string = str(transaction_date_string.split()[0])
            transaction_date = datetime.date.fromisoformat(iso_date_string)

            category_name = transaction[5]
            category_id = categories_dict.get_id(category_name)

            amount = transaction[10]

            this_trans = Transaction(account_id, transaction_date, category_id, amount)

            this_trans.cleared = transaction[9]
            this_trans.number = transaction[2]
            this_trans.tag = transaction[6]

            this_trans.description = transaction[3]
            if this_trans.description:
                previous_description = this_trans.description
            else:
                this_trans.description = previous_description



            this_trans.memo = transaction[4]
            this_trans.tax_item = transaction[8]

            transaction_id = transactions.get_transaction_id(self.db_conn, account_id, transaction_date, category_id, amount)

            if transaction_id is None:
                trans_tab.insert_transaction(this_trans)
                self.rpt.write(f"    {account_name}, {transaction_date}, {category_name}, {amount}\n")
            else:
                transactions.update_transaction(self.db_conn, transaction_id, transaction)